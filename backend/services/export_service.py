import io
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_BG = "FF000000"
_HEADER_FONT = "FFFFFFFF"

_FONT_NAME = "Calibri"

_COL_WIDTHS = [50.88, 15.0, 14.38, 22.25, 13.75, 13.25, 15.5, 15.0]
_MONTH_COL_WIDTH = 13.13

_PT_MONTHS = [
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]

_FIXED_NEGATIVAS = [
    ("GERAL", None, True),
    (
        "Convênios",
        (
            "amil, bradesco saude, unimed, sulamerica, porto seguro, cassi, geap, "
            "plano de saude, convenio, aceita convenio, reembolso (se não trabalhar com reembolso)."
        ),
        False,
    ),
    (
        "Geral/Financeiro",
        (
            "[gratis], [gratuito], [online], [cotação], [preço], [quanto custa] "
            "(se não quiser atrair diretamente por preço), [barato], [desconto] "
            "(à exceção dos convênios), [promocao], [convênio] [publico] [SUS]"
        ),
        False,
    ),
    ("Tipo de Serviço (médico não odontológico)", None, True),
    (
        "Aparelhos/Venda/Conserto",
        "[aparelhos], [maquinas], [equipamentos], [comprar], [venda], [alugar], [conserto], [manutencao], [pecas]",
        False,
    ),
    (
        "Cursos/Profissional",
        "[curso], [faculdade], [graduação], [pos], [especialização], [profissionais], [aprender], [livro], [vaga], [emprego], [carreira], [salario]",
        False,
    ),
    (
        "Geral irrelevante",
        "[fotos], [imagens], [wiki], [youtube], [filme], [serie], [download], [pdf], [videos], [blog], [forum], [artigo], [noticias]",
        False,
    ),
]


def _study_month_year(study: dict) -> tuple[str, int]:
    raw = study.get("created_at")
    dt: datetime | None = None
    if isinstance(raw, str):
        try:
            dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        except ValueError:
            dt = None
    if dt is None:
        dt = datetime.utcnow()
    return _PT_MONTHS[dt.month - 1], dt.year


def _format_location(study: dict) -> str:
    locations = study.get("locations") or []
    if locations:
        first = str(locations[0]).strip()
        cleaned = re.sub(r"^(city|state|country):", "", first, flags=re.IGNORECASE)
        cleaned = cleaned.split("#")[0].strip()
        if cleaned:
            return f"Local: {cleaned.title()}"
    country = study.get("country") or ""
    return f"Local: {country.upper()}" if country else "Local: -"


def _seeds_label(seeds: list[str]) -> str:
    joined = "; ".join(seed.strip() for seed in seeds if seed and seed.strip())
    return f"Palavras-chaves usadas (limite 10):  {joined}"


def _month_columns(reference_month: str, reference_year: int) -> list[str]:
    idx = _PT_MONTHS.index(reference_month)
    cols: list[str] = []
    year = reference_year
    month_idx = idx
    sequence: list[tuple[str, int]] = []
    for _ in range(12):
        sequence.append((_PT_MONTHS[month_idx], year))
        month_idx -= 1
        if month_idx < 0:
            month_idx = 11
            year -= 1
    for name, yr in reversed(sequence):
        cols.append(f"Searches: \n{name} {yr}")
    return cols


def _to_int_or_none(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str):
        cleaned = value.strip().replace(",", "").replace(".", "")
        return int(cleaned) if cleaned.isdigit() else None
    return None


def _yoy_header_label(month_name: str, year: int) -> str:
    short_curr = f"{month_name[:5].lower()}/{str(year)[-2:]}"
    short_prev = f"{month_name[:5].lower()}/{str(year - 1)[-2:]}"
    return (
        "Mudança em relação ao mesmo mês do ano anterior \n"
        f"({short_prev} e {short_curr})"
    )


def _write_meta_and_headers(
    ws,
    seeds: list[str],
    location_label: str,
    month_name: str,
    year: int,
    month_columns: list[str],
) -> None:
    meta_font = Font(name=_FONT_NAME, size=10)
    ws.cell(row=1, column=1, value=f"{month_name} de {year}").font = meta_font
    ws.cell(row=1, column=2, value=location_label).font = meta_font
    ws.cell(row=1, column=4, value=_seeds_label(seeds)).font = meta_font
    ws.row_dimensions[1].height = 18

    headers = [
        "Palavra-Chave",
        "Média de Pesquisas",
        "Mudança em três meses",
        _yoy_header_label(month_name, year),
        "Concorrência",
        "Grau de concorrência",
        "Menores valores para aparecer no topo da pesquisa",
        "Maiores valores para aparecer no topo da pesquisa",
    ] + month_columns

    header_fill = PatternFill("solid", fgColor=_HEADER_BG)
    header_font = Font(name=_FONT_NAME, bold=True, color=_HEADER_FONT, size=10)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 46

    for col_idx in range(1, len(_COL_WIDTHS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = _COL_WIDTHS[col_idx - 1]
    for col_idx in range(len(_COL_WIDTHS) + 1, len(_COL_WIDTHS) + 1 + len(month_columns)):
        ws.column_dimensions[get_column_letter(col_idx)].width = _MONTH_COL_WIDTH

    ws.freeze_panes = "B3"


def _write_data_rows(ws, items: list[dict], month_columns: list[str]) -> None:
    body_font = Font(name=_FONT_NAME, size=10)
    right = Alignment(horizontal="right")
    left = Alignment(horizontal="left")

    monthly_keys = [col.replace("Searches: \n", "Searches: ") for col in month_columns]

    for row_offset, item in enumerate(items, start=3):
        monthly = item.get("searches_mensais") or {}
        media = item.get("media_pesquisas")
        media_int = _to_int_or_none(media)
        change_3m = item.get("mudanca_tres_meses") or "--"
        change_yoy = item.get("mudanca_ano_anterior") or "--"
        competition = item.get("concorrencia") or "Desconhecido"
        comp_index = item.get("grau_concorrencia")
        low_bid = item.get("menor_lance_topo")
        high_bid = item.get("maior_lance_topo")

        row_data = [
            item.get("name", ""),
            media_int,
            change_3m,
            change_yoy,
            competition,
            comp_index,
            low_bid,
            high_bid,
        ]
        for key in monthly_keys:
            row_data.append(_to_int_or_none(monthly.get(key)))

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_offset, column=col_idx, value=value)
            cell.font = body_font
            if col_idx == 1:
                cell.alignment = left
            else:
                cell.alignment = right
            if col_idx == 2:
                cell.number_format = "#,##0;-#,##0;"
            elif col_idx == 6:
                cell.number_format = "0;-0;"
            elif col_idx in (7, 8):
                cell.number_format = "#,##0.00;-#,##0.00;"
            elif col_idx >= 9:
                cell.number_format = "#,##0;-#,##0;"


def _write_sheet(
    ws,
    items: list[dict],
    seeds: list[str],
    location_label: str,
    month_name: str,
    year: int,
) -> None:
    month_columns = _month_columns(month_name, year)
    _write_meta_and_headers(ws, seeds, location_label, month_name, year, month_columns)
    _write_data_rows(ws, items, month_columns)


def _write_negativas(ws) -> None:
    body_font = Font(name=_FONT_NAME, size=10)
    bold_font = Font(name=_FONT_NAME, size=10, bold=True)
    header_fill = PatternFill("solid", fgColor=_HEADER_BG)
    header_font = Font(name=_FONT_NAME, size=10, bold=True, color=_HEADER_FONT)

    for col_idx, header in enumerate(["Categoria", "Palavras Negativas"], start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 110
    ws.freeze_panes = "A2"

    row = 2
    for label, value, is_section in _FIXED_NEGATIVAS:
        cat_cell = ws.cell(row=row, column=1, value=label)
        cat_cell.font = bold_font if is_section else body_font
        cat_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        if value is not None:
            val_cell = ws.cell(row=row, column=2, value=value)
            val_cell.font = body_font
            val_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        row += 1


_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


def _safe_sheet_title(name: str, used: set[str]) -> str:
    cleaned = _INVALID_SHEET_CHARS.sub(" ", name).strip() or "Aba"
    base = cleaned[:31]
    candidate = base
    counter = 2
    while candidate in used:
        suffix = f" ({counter})"
        candidate = (base[: 31 - len(suffix)] + suffix)
        counter += 1
    used.add(candidate)
    return candidate


def _format_location_from_loc_string(loc: str) -> str:
    cleaned = re.sub(r"^(city|state|country):", "", loc.strip(), flags=re.IGNORECASE)
    cleaned = cleaned.split("#")[0].strip()
    return f"Local: {cleaned.title()}" if cleaned else "Local: -"


def _generate_xlsx_multi(study: dict) -> bytes:
    wb = Workbook()
    used_titles: set[str] = set()
    month_name, year = _study_month_year(study)

    default_locations = study.get("default_locations") or []
    default_location_label = (
        _format_location_from_loc_string(default_locations[0])
        if default_locations
        else f"Local: {(study.get('country') or '').upper() or '-'}"
    )

    tabs = study.get("tabs") or []
    first = True
    for tab in tabs:
        name = tab.get("name") or "Aba"
        seeds = tab.get("seeds") or []
        items = tab.get("items") or []
        tab_locations = tab.get("locations") or default_locations
        location_label = (
            _format_location_from_loc_string(tab_locations[0])
            if tab_locations
            else default_location_label
        )

        if first:
            ws = wb.active
            ws.title = _safe_sheet_title(name, used_titles)
            first = False
        else:
            ws = wb.create_sheet(title=_safe_sheet_title(name, used_titles))

        _write_sheet(ws, items, seeds, location_label, month_name, year)

    if first:
        ws = wb.active
        ws.title = _safe_sheet_title("Geral", used_titles)
        _write_sheet(ws, [], [], default_location_label, month_name, year)

    ws_neg = wb.create_sheet(title=_safe_sheet_title("Palavras Negativas", used_titles))
    _write_negativas(ws_neg)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_xlsx(study: dict) -> bytes:
    if isinstance(study.get("tabs"), list):
        return _generate_xlsx_multi(study)

    wb = Workbook()
    used_titles: set[str] = set()

    seeds = list(study.get("seed_keywords") or [])
    location_label = _format_location(study)
    month_name, year = _study_month_year(study)

    ws_general = wb.active
    ws_general.title = _safe_sheet_title("Geral", used_titles)
    _write_sheet(ws_general, study.get("general", []), seeds, location_label, month_name, year)

    categories: dict[str, list[dict]] = study.get("categories", {}) or {}
    for cat_name, items in categories.items():
        ws = wb.create_sheet(title=_safe_sheet_title(cat_name, used_titles))
        _write_sheet(ws, items, seeds, location_label, month_name, year)

    ws_neg = wb.create_sheet(title=_safe_sheet_title("Palavras Negativas", used_titles))
    _write_negativas(ws_neg)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
